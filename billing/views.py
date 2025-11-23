from django.shortcuts import render, redirect, get_object_or_404
from .models import Invoice, InvoiceItem, Product
from .forms import InvoiceForm
from accounts.decorators import role_required
from store.models import StoreProduct
from django.contrib import messages
from django.utils import timezone
from billing.models import CustomerInvoice, TransactionInvoice
from store.models import Product, Store
from accounts.models import UserProfile
from .models import Cart, CartItem, TransactionInvoice, CustomerInvoice
from .utils import get_total_transaction_value
from io import BytesIO
import base64
from django.urls import reverse

from decimal import Decimal



import uuid
# Example of how cart_id might be set in the session
def cart_view(request):
    cart_id = request.session.get('cart_id')
    if not cart_id:
        cart_id = str(uuid.uuid4())  # Generate a new cart_id if not set
        request.session['cart_id'] = cart_id  # Store cart_id in session

    # Now you can use cart_id in your queries




# billing/views.py

import json
import base64
import qrcode
from io import BytesIO
from decimal import Decimal

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.urls import reverse
from django.contrib.auth.decorators import login_required

from .models import CustomerInvoice, TransactionInvoice
from store.models import TaxAndDiscount, Product


@login_required
def invoice_receipt(request, invoice_id):
    # 🔁 Get the return URL (default to sales_view)
    next_url = request.GET.get('next', reverse('billing:sales_view'))

    customer_invoice = get_object_or_404(CustomerInvoice, id=invoice_id)
    cart_items = TransactionInvoice.objects.filter(customer_invoice=customer_invoice)

    if not cart_items.exists():
        messages.error(request, "No items found for this invoice!")
        return redirect('billing:sales_view')

    payment_method_display = customer_invoice.get_payment_method_display()
    subtotal = sum(item.subtotal for item in cart_items)
    total_discount = sum(item.discount for item in cart_items)
    total_tax = sum(item.tax for item in cart_items)
    final_total = customer_invoice.final_total or (subtotal - total_discount + total_tax)
    amount_paid = customer_invoice.amount_paid or Decimal('0.00')
    change = amount_paid - final_total

    invoice_url = request.build_absolute_uri(
        reverse('billing:invoice_receipt', args=[invoice_id])
    )
    qr_img = qrcode.make(invoice_url)
    buffer = BytesIO()
    qr_img.save(buffer, format="PNG")
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()

    store = customer_invoice.store
    store_info = {
        'store_name': store.name,
        'location': store.location,
        'manager_name': store.manager,
        'manager_contact': store.manager_contact,
    }

    context = {
        'customer_invoice': customer_invoice,
        'cart_items': cart_items,
        'payment_method_display': payment_method_display,
        'subtotal': subtotal.quantize(Decimal('0.01')),
        'total_discount': total_discount.quantize(Decimal('0.01')),
        'total_tax': total_tax.quantize(Decimal('0.01')),
        'total': Decimal(final_total).quantize(Decimal('0.01')),
        'amount_paid': amount_paid.quantize(Decimal('0.01')),
        'change': change.quantize(Decimal('0.01')),
        'store_info': store_info,
        'user': request.user,
        'qr_code_base64': qr_code_base64,
        'next_url': next_url,  # ✅ Pass to template
    }

    return render(request, 'billing/invoice_receipt.html', context)


# billing/views.py
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from .models import CustomerInvoice
from .printers import open_cash_drawer  # import your win32print drawer function
from django.shortcuts import get_object_or_404

@login_required
@csrf_exempt
def open_drawer_view(request, invoice_id):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method"}, status=405)

    # Validate invoice exists
    get_object_or_404(CustomerInvoice, id=invoice_id)

    # Trigger cash drawer
    drawer_opened = open_cash_drawer()

    return JsonResponse({"drawer_opened": drawer_opened})



from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from store.models import Product
from billing.models import CustomerInvoice, Cart, TransactionInvoice
from accounts.models import UserProfile
import random
import string
import json

# Utility function to generate unique invoice number
def generate_invoice_number():
    return "INV" + ''.join(random.choices(string.digits, k=6))

@login_required
def sales_view(request):
    if not request.user.is_authenticated:
        return redirect('login')

    # Get user role and store
    try:
        user_profile = request.user.userprofile
        role = user_profile.role
        user_store = user_profile.store
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile does not exist.")
        return redirect('dashboard:index')

    if not user_store:
        messages.error(request, "You are not assigned to any store.")
        return redirect('dashboard:index')

    if not role:
        messages.error(request, "User role is not assigned. Please contact admin.")
        return redirect('dashboard:index')

    # Get or create the last customer invoice for this user & store or create a new one if none
    customer_invoice = CustomerInvoice.objects.filter(user=request.user, store=user_store).last()
    if customer_invoice is None:
        customer_invoice = CustomerInvoice.objects.create(
            invoice_number=generate_invoice_number(),
            total_amount=0,
            discount=0,
            tax=0,
            final_total=0,
            user=request.user,
            store=user_store
        )

    invoice_id = customer_invoice.id

    # Filter products based on user role
    if role == 'cashier' or role == 'admin':
        products = Product.objects.filter(store=user_store, status=True)
    elif role == 'staff':
        products = Product.objects.filter(store=user_store, status=True)  # Adjust filter if needed
    else:
        products = Product.objects.none()  # Default empty queryset for unknown roles


    # Identify low stock products
    low_stock_products = [product for product in products if product.is_stock_low]

    # Get or create cart
    cart_id = request.session.get('cart_id', None)
    if cart_id is None:
        cart = Cart.objects.create(user=request.user, store=user_store)
        request.session['cart_id'] = cart.id
        cart_id = cart.id
    else:
        try:
            cart = Cart.objects.get(id=cart_id)
        except Cart.DoesNotExist:
            cart = Cart.objects.create(user=request.user, store=user_store)
            request.session['cart_id'] = cart.id
            cart_id = cart.id

    # Clear cart items if requested
    if 'reset_cart' in request.GET:
        cart.cart_items.all().delete()

    cart_items_qs = cart.cart_items.all()

    # Calculate total amount for cart items
    total_amount = sum(item.quantity * item.product.selling_price for item in cart_items_qs)

    # Sync cart contents into session as JSON
    cart_items = [{'id': item.product.id, 'quantity': item.quantity} for item in cart_items_qs]
    request.session['cart_data'] = json.dumps(cart_items)
    request.session.modified = True

    return render(request, 'billing/sales.html', {
        'invoice_id': invoice_id,
        'products': products,
        'low_stock_products': low_stock_products,
        'cart': cart,
        'cart_items': cart_items_qs,
        'total_amount': total_amount,
    })




def serialize_low_stock(products):
    # Customize the fields you want to send
    return [
        {
            'id': p.id,
            'name': p.name,
            'quantity': p.quantity,
        } for p in products
    ]





from .models import CustomerInvoice

def invoice_success(request):
    # You can query the last created invoice or whatever logic you need
    invoice = CustomerInvoice.objects.last()  # Example logic
    return render(request, 'store/invoice_success.html', {'invoice': invoice})


def create_invoice(request):
    # Add your logic for creating an invoice here
    return render(request, 'billing/create_invoice.html')

# billing/views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import Product, Cart

def edit_quantity(request, cart_item_id):
    cart_item = get_object_or_404(Cart, id=cart_item_id)

    if request.method == 'POST':
        new_quantity = request.POST.get('quantity')

        # Validate the new quantity
        if new_quantity and int(new_quantity) > 0:
            cart_item.quantity = int(new_quantity)
            cart_item.save()
            return redirect('billing:sales')  # Redirect to the sales page after update

    return render(request, 'billing/edit_quantity.html', {
        'cart_item': cart_item,
    })

from django.http import JsonResponse
from .models import Cart


from django.http import JsonResponse
from .models import Cart
def delete_item(request, cart_item_id):
    print(f"Attempting to delete cart item with ID: {cart_item_id}")  # Log this to see if it's reached
    try:
        cart_item = Cart.objects.get(id=cart_item_id)
        cart_item.delete()

        # Optionally, update session data
        cart = request.session.get('cart', [])
        cart = [item for item in cart if item['id'] != cart_item_id]  # Remove item from session
        request.session['cart'] = cart

        return JsonResponse({'success': True})
    except Cart.DoesNotExist:
        
        return JsonResponse({'error': 'Cart item not found.'}, status=404)


from django.contrib.auth.decorators import login_required
from billing.models import TransactionInvoice

@login_required
def store_transactions(request):
    # Ensure that the user has a store associated with them (Either via profile or othe menas)
    store = request.user.store
    # Filter transactions based on the user's assigned store
    transactions = TransactionInvoice.objects.filter(store=store)
    return render(request, 'billing/transactions.html', {'transactions': transactions})


# billing/views.py (or where you're processing the sale)

from .models import TransactionInvoice
from store.models import Product  # Assuming this is the model tracking store products

from store.models import Product  # This is your Product model

def update_stock_after_sale(product_id, quantity_sold):
    try:
        # Fetch the product from the store
        product = Product.objects.get(id=product_id)

        # Reduce the stock by the quantity sold
        if product.quantity >= quantity_sold:
            product.quantity -= quantity_sold
            product.save()
        else:
            raise ValueError("Not enough stock available")
    except Product.DoesNotExist:
        raise ValueError("Product not found in the store")


# View for processing the sale (simplified)
# billing/views.py

from django.shortcuts import render, redirect
from .models import Product, TransactionInvoice, CustomerInvoice
from django.contrib.auth.decorators import login_required

@login_required
def process_sale(request):
    if request.method == "POST":
        customer_name = request.POST.get('customer_name')
        product_ids = request.POST.getlist('product_ids')
        quantities = request.POST.getlist('quantities')
        discount = float(request.POST.get('discount', 0))
        tax = float(request.POST.get('tax', 0))
        amount_paid = float(request.POST.get('amount_paid', 0))

        # Calculate final total
        subtotal = 0
        cart_items = []
        for product_id, quantity in zip(product_ids, quantities):
            product = Product.objects.get(id=product_id)
            quantity = int(quantity)
            total_price = product.selling_price * quantity
            discounted_price = total_price - (total_price * discount / 100)
            cart_items.append({
                'product': product,
                'quantity': quantity,
                'total_price': discounted_price
            })
            subtotal += discounted_price

        # Calculate Tax and Final Total
        tax_amount = subtotal * (tax / 100)
        final_total = subtotal + tax_amount

        # Save Customer Invoice
        customer_invoice = CustomerInvoice.objects.create(
            customer_name=customer_name,
            total_amount=final_total,
            discount=discount,
            tax=tax_amount,
            payment_status="Paid"  # Modify based on your payment status
        )

        # Update Stock and Create Transaction Invoices
        for item in cart_items:
            product = item['product']
            if product.quantity >= item['quantity']:
                product.quantity -= item['quantity']
                product.save()

                # ✅ Proportionally distribute tax based on the product's contribution to subtotal
                item_tax = (item['total_price'] / subtotal) * tax_amount if subtotal > 0 else 0

                # Create Transaction Invoice for each item
                TransactionInvoice.objects.create(
                    product=product,
                    quantity=item['quantity'],
                    price=item['total_price'],
                    invoice=customer_invoice,
                    tax=item_tax  # ✅ Save tax for each transaction
                )
            else:
                # Handle out-of-stock scenario
                return render(request, 'billing/sales.html', {
                    'error': 'Not enough stock for product: ' + product.name
                })

        # Redirect to the invoice receipt page
        return redirect('billing:invoice_receipt', invoice_id=customer_invoice.id)

    else:
        # Fetch available products for sale
        products = Product.objects.filter(status=True)  # Only active products
        return render(request, 'billing/sales.html', {'products': products})


# import json
import json
from decimal import Decimal
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from .models import CustomerInvoice, TransactionInvoice, Customer
from store.models import Product, TaxAndDiscount
from .printers import open_cash_drawer


@login_required
@csrf_protect
def generate_invoice(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    try:
        payload = json.loads(request.body)
        print("Payload received:", payload)  # Debug log

        raw_name = payload.get('customer_name', '').strip()
        raw_phone = payload.get('phone_number', '').strip()

        customer_name = raw_name or 'Customer'
        phone_number = raw_phone if raw_phone else None

        print(f"Customer name: '{customer_name}', Phone number: '{phone_number}'")  # Debug log

        cart_items = payload.get('cart', [])
        if not cart_items:
            return JsonResponse({'error': 'Missing cart data'}, status=400)

        # --- STOCK QUANTITY VALIDATION BEFORE CREATING INVOICE ---
        for item in cart_items:
            pid = item.get('product_id') or item.get('id')
            quantity = int(item.get('quantity', 0))
            product = get_object_or_404(Product, id=pid)
            if quantity > product.quantity:
                return JsonResponse({
                    'error': f"Insufficient stock for '{product.name}'. Available: {product.quantity}, requested: {quantity}."
                }, status=400)

        # Proceed with customer get-or-create logic
        if phone_number:
            customer = Customer.objects.filter(name=customer_name, phone_number__isnull=True).first()
            if customer:
                customer.phone_number = phone_number
                customer.save()
            else:
                customer, created = Customer.objects.get_or_create(
                    phone_number=phone_number,
                    defaults={'name': customer_name}
                )
                if not created and customer.name != customer_name:
                    customer.name = customer_name
                    customer.save()
        else:
            customer = Customer.objects.filter(name=customer_name).first()
            if not customer:
                customer = Customer.objects.create(name=customer_name, phone_number=None)

        amount_paid = Decimal(payload.get('amount_paid') or '0')
        payment_method = payload.get('payment_method', 'Unknown')

        td = TaxAndDiscount.objects.first()
        discount_rate = td.discount if td else Decimal('0')
        tax_rate = td.tax if td else Decimal('0')

        store = request.user.store

        invoice = CustomerInvoice.objects.create(
            customer=customer,
            customer_name=customer_name,
            total_amount=Decimal('0.00'),
            tax=Decimal('0.00'),
            discount=Decimal('0.00'),
            final_total=Decimal('0.00'),
            amount_paid=amount_paid,
            change=Decimal('0.00'),
            payment_method=payment_method,
            user=request.user,
            store=store
        )

        total_amount = Decimal('0.00')
        total_discount = Decimal('0.00')
        total_tax = Decimal('0.00')

        for item in cart_items:
            pid = item.get('product_id') or item.get('id')
            quantity = int(item.get('quantity', 0))
            price = Decimal(str(item.get('price', '0')))
            product = get_object_or_404(Product, id=pid)
            discounted_price = price * (Decimal('1') - discount_rate / Decimal('100'))
            line_subtotal = (discounted_price * quantity).quantize(Decimal('0.01'))
            total_amount += line_subtotal

            tx = TransactionInvoice.objects.create(
                customer_invoice=invoice,
                product=product,
                quantity=quantity,
                price=price,
                subtotal=line_subtotal,
                discount=Decimal('0.00'),
                tax=Decimal('0.00'),
                prorated_discount=Decimal('0.00'),
                prorated_tax=Decimal('0.00'),
                adjusted_final_price=line_subtotal,
                store=store,
                cart_id=str(invoice.id),
                user=request.user
            )
            total_discount += tx.discount
            total_tax += tx.tax

        invoice.total_amount = total_amount
        invoice.discount = total_discount
        invoice.tax = total_tax
        invoice.final_total = (total_amount + total_tax - total_discount).quantize(Decimal('0.01'))

        change = (amount_paid - invoice.final_total).quantize(Decimal('0.01'))
        invoice.change = change

        invoice.save()

        

        return JsonResponse({'invoice_id': invoice.id})


    except Exception as e:
        print(f"Error in generate_invoice: {e}")
        return JsonResponse({'error': str(e)}, status=500)




from django.shortcuts import redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt  # Optional, only if you want to disable CSRF temporarily
from django.contrib import messages
from store.models import Product
from billing.models import Cart, TransactionInvoice

from django.core.exceptions import ValidationError

from billing.models import Cart, CartItem  # Make sure to import CartItem

def add_to_cart(request):
    if request.method == 'POST':
        cart_id = request.session.get('cart_id')

        if not cart_id:
            cart = Cart.objects.create(user=request.user, store=request.user.userprofile.store)
            request.session['cart_id'] = cart.id
            request.session.modified = True
        else:
            try:
                cart = Cart.objects.get(id=cart_id)
            except Cart.DoesNotExist:
                cart = Cart.objects.create(user=request.user, store=request.user.userprofile.store)
                request.session['cart_id'] = cart.id
                request.session.modified = True

        product_id = request.POST.get('product_id')
        quantity = int(request.POST.get('quantity', 1))

        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': f"Product with ID {product_id} not found!"}, status=404)

        # Check if CartItem exists for product in this cart
        cart_item, created = CartItem.objects.get_or_create(cart=cart, product=product)

        if not created:
            cart_item.quantity += quantity  # Increase quantity if already exists
        else:
            cart_item.quantity = quantity

        cart_item.save()

        return JsonResponse({'status': 'success', 'message': f"{product.name} added to cart successfully!"})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)






def generate_new_cart_id():
    # Generate a unique cart ID (for example, using UUID)
    import uuid
    return str(uuid.uuid4())

import uuid
# Example of how cart_id might be set in the session
def cart_view(request):
    cart_id = request.session.get('cart_id')
    if not cart_id:
        cart_id = str(uuid.uuid4())  # Generate a new cart_id if not set
        request.session['cart_id'] = cart_id  # Store cart_id in session

    # Now you can use cart_id in your queries

# billing/views.py

from .forms import CartForm
from .models import Cart

@login_required
def add_cart(request):
    if request.method == "POST":
        form = CartForm(request.POST)
        if form.is_valid():
            cart = form.save(commit=False)
            cart.user = request.user  # Assign the logged-in user
            cart.save()
            messages.success(request, "Cart added successfully!")
            return redirect('billing:add_cart')  # Adjust to your desired redirect
    else:
        form = CartForm()

    return render(request, 'billing/add_cart.html', {'form': form})


from django.utils import timezone
from .models import Cart

def create_cart(request, store_id):
    user = request.user  # Get the logged-in user (cashier)
    
    # Create a new cart
    cart = Cart.objects.create(
        created_at=timezone.now(),
        is_paid=False,  # Initial status of the cart
        store_id=store_id,
        user=user  # Associate the cart with the cashier
    )
    
    # Save the cart ID to the session
    request.session['cart_id'] = cart.id
    print(f"New cart created with ID: {cart.id} and saved in the session.")

    return cart.id  # Return cart_id

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .models import Cart, Product

@csrf_exempt  # Add this decorator if you're using a non-CSRF-checked endpoint
@require_POST  # Ensure the view only accepts POST requests
def save_cart(request):
    try:
        cart_data = json.loads(request.body)  # Parse the JSON from the request
        cart = cart_data.get('cart')
        cart_id = cart_data.get('cart_id')

        # You can now process the cart data here, e.g., saving to the database
        if cart_id:
            cart_instance = Cart.objects.get(id=cart_id)
            # You can filter or process the cart items based on cart_id

            # Example of processing the cart (saving transaction, etc.)
            for item in cart:
                product = Product.objects.get(id=item['id'])
                # Add logic to handle saving items in the cart
                # ...

            return JsonResponse({'success': True, 'message': 'Cart saved successfully.'})

        return JsonResponse({'success': False, 'error': 'Cart ID missing.'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})



#=======================================================
# SAVING CREATED CART_ID FROM JAVASCRIPT
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

@csrf_exempt
def save_cart_id(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            cart_id = int(data.get('cart_id', 0))  # Ensure cart_id is converted to an integer
            request.session['cart_id'] = cart_id
            return JsonResponse({'message': 'Cart ID saved successfully!'})
        except (ValueError, TypeError) as e:
            return JsonResponse({'error': 'Invalid cart ID format.'}, status=400)
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


from django.shortcuts import render

def cart_view(request):
    cart_id = request.session.get('cart_id', None)
    return render(request, 'billing/cart_template.html', {'cart_id': cart_id})

import random
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def clear_cart(request):
    # Clear the cart and cart_id from the session
    request.session.pop('cart', None)  # Removes cart from session if it exists
    request.session.pop('cart_id', None)  # Removes cart_id from session if it exists
    
    # Generate a new cart_id and store it
    cart_id = str(random.randint(1000000000, 9999999999))  # New 10-digit cart_id
    request.session['cart_id'] = cart_id
    
    return JsonResponse({'status': 'success', 'message': 'Cart cleared and new cart ID generated.', 'cart_id': cart_id})



def all_transactions(request):
    filter_field = request.GET.get('filter_field')
    filter_value = request.GET.get('filter_value')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    transactions = TransactionInvoice.objects.filter(
    customer_invoice__is_void=False
    ).order_by('-created_at')


    # Apply filter_field and filter_value
    if filter_field and filter_value:
        if filter_field == 'product':
            transactions = transactions.filter(product__name__icontains=filter_value)
        elif filter_field == 'store_name':
            transactions = transactions.filter(store__name__icontains=filter_value)

    # Apply date filters
    if start_date:
        transactions = transactions.filter(created_at__gte=start_date)
    if end_date:
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        end_date_obj = end_date_obj + timedelta(days=1) - timedelta(seconds=1)
        transactions = transactions.filter(created_at__lte=end_date_obj)

    totals = transactions.aggregate(
        total_sales=Sum('subtotal'),
        total_quantity=Sum('quantity')
    )
    total_sales = totals['total_sales'] or 0
    total_quantity = totals['total_quantity'] or 0

    paginator = Paginator(transactions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'billing/transactions.html', {
        'transactions': page_obj,
        'total_sales': total_sales,
        'total_quantity': total_quantity,
    })



from django.shortcuts import render
from django.db.models import Q, Sum
from datetime import datetime, timedelta
from django.core.paginator import Paginator
from billing.models import TransactionInvoice
from store.models import TaxAndDiscount
from django.http import JsonResponse
from django.utils.dateparse import parse_date
import logging


# ====== Transaction List View ======
def transactions_list(request):
    # Fetch transactions with necessary related data, excluding voided ones
    transactions = (
        TransactionInvoice.objects
        .select_related('customer_invoice', 'product', 'store')
        .exclude(is_void=True)  # ✅ exclude voided transactions
    )

    # Fetch global tax and discount settings (assumes only one global config)
    global_tax_discount = TaxAndDiscount.objects.first()
    global_discount = global_tax_discount.discount if global_tax_discount else 0
    global_tax = global_tax_discount.tax if global_tax_discount else 0

    # Calculate adjusted total sales after discount and tax
    total_sales = 0
    for transaction in transactions:
        product_discount = transaction.discount or 0
        product_tax = transaction.tax or 0

        total_discount = product_discount + global_discount
        total_tax = product_tax + global_tax

        transaction.prorated_discount = (transaction.subtotal * total_discount) / 100
        transaction.prorated_tax = (transaction.subtotal * total_tax) / 100
        transaction.adjusted_final_price = (
            transaction.subtotal - transaction.prorated_discount + transaction.prorated_tax
        )

    # Calculate total adjusted sales correctly
    total_sales = sum(
        getattr(t, 'adjusted_final_price', 0) for t in transactions
    )
    total_sales = round(total_sales, 2)

    # Handle search and filter queries
    search_query = request.GET.get('search', '')
    filter_period = request.GET.get('filter', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    filter_field = request.GET.get('filter_field', '')
    filter_value = request.GET.get('filter_value', '')

    # Apply search query filter
    if search_query:
        transactions = transactions.filter(
            Q(store__name__icontains=search_query) |
            Q(customer_invoice__customer_name__icontains=search_query) |
            Q(product__name__icontains=search_query)
        )

    # Apply specific field filters
    if filter_field and filter_value:
        if filter_field == 'store':
            transactions = transactions.filter(store__name__icontains=filter_value)
        elif filter_field == 'invoice_number':
            transactions = transactions.filter(customer_invoice__invoice_number__icontains=filter_value)
        elif filter_field == 'customer_name':
            transactions = transactions.filter(customer_invoice__customer_name__icontains=filter_value)
        elif filter_field == 'payment_method':
            transactions = transactions.filter(customer_invoice__payment_method__icontains=filter_value)

    # Apply period filters
    if filter_period:
        today = datetime.today()
        if filter_period == 'day':
            transactions = transactions.filter(created_at__date=today.date())
        elif filter_period == 'week':
            start_of_week = today - timedelta(days=today.weekday())
            transactions = transactions.filter(created_at__date__gte=start_of_week)
        elif filter_period == 'month':
            transactions = transactions.filter(created_at__month=today.month)
        elif filter_period == 'year':
            transactions = transactions.filter(created_at__year=today.year)

    # Apply date range filter
    try:
        if start_date and end_date:
            start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d').date()
            transactions = transactions.filter(created_at__date__range=[start_date_parsed, end_date_parsed])
    except ValueError:
        pass

    # Order and paginate
    transactions = transactions.order_by('-created_at')
    paginator = Paginator(transactions, 10)
    page_number = request.GET.get('page')
    transactions_page = paginator.get_page(page_number)

    return render(request, 'billing/transactions_list.html', {
        'transactions': transactions_page,
        'total_sales': total_sales,
    })


# ====== Filter Transactions (AJAX) ======
logger = logging.getLogger(__name__)

def filter_transactions(request):
    try:
        filter_field = request.GET.get('filter_field')
        filter_value = request.GET.get('filter_value')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Exclude voided transactions
        transactions = TransactionInvoice.objects.exclude(is_void=True)

        # Apply date range
        if start_date and end_date:
            start_date = parse_date(start_date)
            end_date = parse_date(end_date)
            if start_date and end_date:
                transactions = transactions.filter(created_at__date__range=[start_date, end_date])

        # Apply field-based filtering
        if filter_field and filter_value:
            filter_criteria = {}
            if filter_field == 'store':
                filter_criteria = {'store__name__icontains': filter_value}
            elif filter_field == 'invoice_number':
                filter_criteria = {'cart_id__icontains': filter_value}
            elif filter_field == 'customer_name':
                filter_criteria = {'customer_name__icontains': filter_value}
            elif filter_field == 'payment_method':
                filter_criteria = {'payment_method__icontains': filter_value}
            elif filter_field == 'cashier_name':
                filter_criteria = {'user__username__icontains': filter_value}

            transactions = transactions.filter(**filter_criteria)

        # Calculate totals
        subtotal_sum = transactions.aggregate(Sum('subtotal'))['subtotal__sum'] or 0
        tax_sum = transactions.aggregate(Sum('tax'))['tax__sum'] or 0
        total_sales = subtotal_sum + tax_sum

        # Serialize transactions
        transaction_data = [
            {
                'cashier_name': transaction.user.username if transaction.user else 'Unknown',
                'cart_id': transaction.cart_id,
                'customer_name': transaction.customer_name,
                'payment_method': transaction.payment_method,
                'product_name': transaction.product.name,
                'store_name': transaction.store.name,
                'quantity': transaction.quantity,
                'price': transaction.price,
                'discount': transaction.product.discount,
                'tax': transaction.product.product_tax if transaction.product.product_tax is not None else 0.00,
                'subtotal': transaction.subtotal,
                'prorated_discount': getattr(transaction, 'prorated_discount', 0.00),
                'prorated_tax': getattr(transaction, 'prorated_tax', 0.00),
                'adjusted_final_price': getattr(transaction, 'adjusted_final_price', 0.00),
                'created_at': transaction.created_at.strftime('%d-%m-%Y') if transaction.created_at else 'N/A',
            }
            for transaction in transactions
        ]

        return JsonResponse({'transactions': transaction_data, 'total_sales': total_sales})

    except Exception as e:
        logger.error(f"Error in filter_transactions view: {str(e)}")
        return JsonResponse({'error': 'Internal Server Error'}, status=500)



from django.shortcuts import render
from .utils import get_total_transaction_value

def total_transaction_value(request):
    total_value = get_total_transaction_value()
    return render(request, 'billing/transaction_value.html', {'total_value': total_value})


#=====================================================
#  Working on refund



# Set up logging
import logging
from django.http import JsonResponse
from django.shortcuts import render
from .models import TransactionInvoice, CustomerInvoice, Product


def transaction_search(request):
    if request.method == "GET":
        bill_number = request.GET.get('bill_number')
        if bill_number:
            transactions = TransactionInvoice.objects.filter(cart_id__startswith=bill_number)
            if not transactions.exists():
                return JsonResponse({"error": "No transactions found for this bill number."}, status=404)

            cart_items = [
                {
                    "transaction_id": transaction.id,
                    "name": transaction.product.name,
                    "quantity": transaction.quantity,
                    "price": transaction.price,
                    "subtotal": transaction.subtotal,
                }
                for transaction in transactions
            ]
            return JsonResponse({"cart_items": cart_items})

        return render(request, 'billing/transaction_search.html')



from django.http import HttpResponse
import openpyxl
from openpyxl.chart import PieChart, Reference
from collections import defaultdict
from .models import TransactionInvoice
from datetime import datetime
from django.db.models import Sum, Count


def set_column_widths(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        worksheet.column_dimensions[column_letter].width = max_length + 2


def export_transactions_to_excel(request):
    # Retrieve parameters
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    filter_field = request.GET.get('filter_field', '').strip()
    filter_value = request.GET.get('filter_value', '').strip()

    # Map frontend filter keys to model fields
    field_map = {
        'store_name': 'store__name',
        'product': 'product__name',
    }

    # ✅ Exclude voided invoices properly
    transactions = TransactionInvoice.objects.filter(is_void=False)
    print(f"Initial (non-voided) transactions count: {transactions.count()}")

    # Apply date filtering safely with full datetime boundaries
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            transactions = transactions.filter(created_at__gte=start)
            print(f"After start_date filter ({start_date}): {transactions.count()}")
        except Exception as e:
            print(f"Invalid start_date format: {start_date} | Error: {e}")

    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            # Set end time to end of the day for inclusive filter
            end = end.replace(hour=23, minute=59, second=59)
            transactions = transactions.filter(created_at__lte=end)
            print(f"After end_date filter ({end_date}): {transactions.count()}")
        except Exception as e:
            print(f"Invalid end_date format: {end_date} | Error: {e}")

    # Apply field filter if valid
    if filter_field and filter_value:
        actual_field = field_map.get(filter_field)
        if actual_field:
            filter_value = filter_value.strip()
            transactions = transactions.filter(**{f"{actual_field}__icontains": filter_value})
            print(f"After field filter '{actual_field}' contains '{filter_value}': {transactions.count()}")
        else:
            print(f"Filter field '{filter_field}' not recognized; skipping filter.")

    count = transactions.count()
    print(f"Final transactions count after filtering: {count}")

    # Handle empty result case
    if count == 0:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        headers = ["Product", "Store", "Quantity", "Price", "Discount", "Tax",
                   "Subtotal", "Adjusted Final Price", "Date"]
        ws.append(headers)
        ws.append(["No data found for the applied filters."])
        set_column_widths(ws)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'
        wb.save(response)
        return response

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "Product", "Store", "Quantity", "Price", "Discount", "Tax",
        "Subtotal", "Adjusted Final Price", "Date"
    ]
    ws.append(headers)

    product_summary = defaultdict(lambda: {
        'quantity': 0, 'total_price': 0.0, 'total_unit_price': 0.0, 'count': 0
    })

    for idx, transaction in enumerate(transactions, start=2):
        ws[f'A{idx}'] = transaction.product.name
        ws[f'B{idx}'] = transaction.store.name
        ws[f'C{idx}'] = transaction.quantity
        ws[f'D{idx}'] = float(transaction.price)
        ws[f'E{idx}'] = float(transaction.discount)
        ws[f'F{idx}'] = float(transaction.tax)
        ws[f'G{idx}'] = float(transaction.subtotal)
        ws[f'H{idx}'] = float(transaction.adjusted_final_price)
        ws[f'I{idx}'] = transaction.created_at.strftime('%d-%m-%Y') if transaction.created_at else "N/A"

        pname = transaction.product.name
        product_summary[pname]['quantity'] += transaction.quantity
        product_summary[pname]['total_price'] += transaction.quantity * float(transaction.price)
        product_summary[pname]['total_unit_price'] += float(transaction.price)
        product_summary[pname]['count'] += 1

    summary = transactions.aggregate(
        total_quantity=Sum('quantity'),
        total_subtotal=Sum('subtotal'),
        total_sales=Sum('adjusted_final_price'),
        total_invoices=Count('cart_id', distinct=True)
    )

    total_tax = sum((t.tax or 0) for t in transactions)
    total_discount = sum((t.discount or 0) for t in transactions)

    ws.append([])
    ws.append([
        "", "", "TOTALS", "",
        total_discount,
        total_tax,
        summary['total_subtotal'] or 0,
        summary['total_sales'] or 0,
        "",
    ])
    ws.append(["", "", "Total Quantity", summary['total_quantity'] or 0])
    ws.append(["", "", "Total Invoices", summary['total_invoices'] or 0])

    set_column_widths(ws)

    # Product Quantity Summary Sheet
    summary_ws = wb.create_sheet(title="Product Quantity Summary")
    summary_ws.append(["Product", "Total Quantity", "Avg Price", "Total Price (₦)"])

    row_idx = 1
    for row_idx, (product_name, data) in enumerate(product_summary.items(), start=2):
        avg_price = data['total_unit_price'] / data['count'] if data['count'] else 0
        summary_ws[f"A{row_idx}"] = product_name
        summary_ws[f"B{row_idx}"] = data['quantity']
        summary_ws[f"C{row_idx}"] = round(avg_price, 2)
        summary_ws[f"D{row_idx}"] = round(data['total_price'], 2)

    set_column_widths(summary_ws)

    # Pie Chart
    pie = PieChart()
    pie.title = "Quantity Sold per Product"
    labels = Reference(summary_ws, min_col=1, min_row=2, max_row=row_idx)
    data = Reference(summary_ws, min_col=2, min_row=1, max_row=row_idx)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 10
    pie.width = 10
    summary_ws.add_chart(pie, "F2")

    # Return Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'
    wb.save(response)
    return response

# views.py

import qrcode
from io import BytesIO
import base64
from django.shortcuts import render
from .models import TransactionInvoice
import json

def re_print_invoice(request):
    if request.method == "POST":
        data = json.loads(request.body)
        cart_id = data.get('cart_id')
        print(f"Received cart_id: {cart_id}")  # Debug

        if cart_id:
            transactions = TransactionInvoice.objects.filter(cart_id=cart_id)
            print(f"Transactions found: {transactions.count()}")  # Debug

            if transactions.exists():
                # Calculate subtotal, discount, tax, and total
                subtotal = sum(item.subtotal for item in transactions)
                discount = sum(item.discount for item in transactions)
                tax = sum(item.tax for item in transactions)
                total = subtotal - discount + tax

                # Retrieve amount_paid and change from the first transaction
                # Assuming the `amount_paid` and `change` values are the same for all items in the cart
                first_transaction = transactions.first()
                amount_paid = first_transaction.amount_paid if first_transaction else 0
                change = first_transaction.change if first_transaction else 0

                # Generate QR code as done in the invoice_receipt view
                invoice_url = request.build_absolute_uri(reverse('billing:re_print_invoice'))
                local_ip = "192.168.74.238"  # Replace with your actual local IP
                qr = qrcode.make(f"http://{local_ip}:8000/billing/re_print_invoice/{cart_id}")

                buffer = BytesIO()
                qr.save(buffer, format="PNG")
                qr_base64 = base64.b64encode(buffer.getvalue()).decode()

                # Render the template with all the required context
                return render(request, 'billing/re_print_invoice.html', {
                    'cart_items': transactions,
                    'subtotal': subtotal,
                    'discount': discount,
                    'tax': tax,
                    'total': total,
                    'amount_paid': amount_paid,
                    'change': change,
                    'qr_code_base64': qr_base64,  # Pass the QR code to the template
                })
            else:
                return render(request, 'billing/re_print_invoice.html', {'error': 'No transactions found.'})

    return render(request, 'billing/re_print_invoice.html', {'error': 'No cart ID provided.'})



from django.shortcuts import render, redirect

def reset_sales_page(request):
    # Clear cart session
    if 'cart' in request.session:
        del request.session['cart']
    request.session.modified = True

    # Redirect to the sales page
    return redirect('billing:sales_view')  # Ensure 'sales' is the name of your URL pattern for sales.html


import os
import qrcode
import threading
from io import BytesIO
from tempfile import NamedTemporaryFile
from google.cloud import storage
from django.conf import settings

def upload_qr_to_gcs(local_file_path, cloud_file_name):
    """Upload QR code to Google Cloud Storage asynchronously."""
    try:
        storage_client = storage.Client()
        bucket = storage_client.bucket(settings.GS_BUCKET_NAME)
        blob = bucket.blob(cloud_file_name)
        
        # Upload file
        blob.upload_from_filename(local_file_path, content_type='image/png')

        # Make publicly accessible (optional)
        blob.make_public()

        print(f"✅ QR Code uploaded to: {blob.public_url}")
        return blob.public_url
    except Exception as e:
        print(f"❌ Error uploading QR Code: {e}")
        return ""
    

import os
import qrcode
from tempfile import NamedTemporaryFile
from django.urls import reverse

def generate_qr_code(invoice_id, request):
    """Generate QR Code with Invoice URL and store it locally."""
    
    temp_dir = os.path.join(os.getcwd(), "tmp_dir")
    os.makedirs(temp_dir, exist_ok=True)  # Create if it doesn't exist

    # ✅ Step 1: Generate the correct invoice URL
    ngrok_url = "https://3d13-102-176-65-243.ngrok-free.app"  # Replace this with your Ngrok URL
    invoice_url = ngrok_url + reverse('billing:invoice_receipt', args=[invoice_id])

    # ✅ Step 2: Generate QR Code from the invoice URL
    qr = qrcode.make(invoice_url)
    
    # ✅ Step 3: Save QR Code in a temporary file
    with NamedTemporaryFile(delete=False, suffix=".png", dir=temp_dir) as temp_file:
        qr_path = temp_file.name  # Get temp file path
        qr.save(qr_path)  # Save QR code to temp file

    print(f"✅ QR Code cached locally: {qr_path}")

    # Return the path of the locally saved QR code
    return qr_path  # This is the local path of the QR code

def transaction_receipt(request, transaction_id):
    from django.utils.safestring import mark_safe  # For safe URL handling

    # Get the transaction data
    transaction = get_object_or_404(TransactionInvoice, id=transaction_id)

    # Generate the QR code for this transaction
    cart_id = transaction.cart_id
    qr_code_file = generate_qr_code(cart_id)

    # Construct the correct URL for the QR code
    qr_code_url = settings.MEDIA_URL + os.path.basename(qr_code_file) if qr_code_file else ""

    # Debugging: Print values in terminal
    print(f"Transaction ID: {transaction_id}")
    print(f"Cart ID: {cart_id}")
    print(f"QR Code File Path: {qr_code_file}")
    print(f"Generated QR Code URL: {qr_code_url}")

    # Pass the QR code URL to the template
    return render(request, 'billing/invoice_receipt.html', {
        'transaction': transaction,
        'qr_code_url': mark_safe(qr_code_url),  # Prevent escaping of URL
        'media_url': settings.MEDIA_URL
    })



from django.http import JsonResponse
from store.models import Product

def search_billing_product(request):
    query = request.GET.get("q", "").strip()
    products = Product.objects.filter(name__icontains=query)[:10]

    response_data = []
    for product in products:
        discounted_price = product.discounted_price
        final_price = product.taxed_price

        response_data.append({
            "id": product.id,
            "name": product.name,
            "selling_price": str(product.selling_price),
            "discount": str(product.discount),
            "product_tax": str(product.product_tax),
            "discounted_price": f"{discounted_price:.2f}",
            "final_price": f"{final_price:.2f}",  # <- ✅ Send final price for billing
            "quantity": int(product.quantity),
            "expiry_date": product.expiry_date.strftime("%Y-%m-%d") if product.expiry_date else "",
        })

    return JsonResponse(response_data, safe=False)



from django.http import JsonResponse
from store.models import Product

def search_product(request):
    barcode = request.GET.get('barcode', '').strip()
    if barcode:
        try:
            product = Product.objects.get(barcode=barcode)
            product_data = {
                'name': product.name,
                'price': product.selling_price,
                'quantity': product.quantity,
                'discounted_price': product.discounted_price,
                'taxed_price': product.taxed_price,
                'discount': product.discount,
                'tax': product.tax,
            }
            return JsonResponse({'product': product_data})
        except Product.DoesNotExist:
            return JsonResponse({'error': 'Product not found'}, status=404)
    return JsonResponse({'error': 'Invalid barcode'}, status=400)


from django.http import JsonResponse
from django.db.models import Q
from .models import Customer

def customer_list(request):
    query = request.GET.get('q', '')
    customers = Customer.objects.filter(
        Q(name__icontains=query) | Q(phone_number__icontains=query)
    ).order_by('-id') if query else Customer.objects.all().order_by('-id')

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        data = [
            {
                'id': c.id,
                'name': c.name,
                'phone_number': c.phone_number or '—'
            } for c in customers
        ]
        return JsonResponse({'customers': data})

    return render(request, 'billing/customer_list.html', {'customers': customers})


#===================================================================
from django.shortcuts import render
from .models import CustomerInvoice
from django.core.paginator import Paginator
from datetime import datetime
from django.db.models import Sum
from django.utils import timezone

def invoice_list(request):
    # ✅ Exclude voided invoices here
    invoices = CustomerInvoice.objects.select_related('store', 'user') \
        .filter(is_void=False) \
        .order_by('-created_at')

    # Get filters
    query = request.GET.get('query')
    field = request.GET.get('field')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Field-based search
    if query and field:
        if field == 'invoice_number':
            invoices = invoices.filter(invoice_number__icontains=query)
        elif field == 'user':
            invoices = invoices.filter(user__username__icontains=query)
        elif field == 'store':
            invoices = invoices.filter(store__name__icontains=query)
        elif field == 'customer_name':
            invoices = invoices.filter(customer_name__icontains=query)
        elif field == 'payment_method':
            invoices = invoices.filter(payment_method__icontains=query)

    # Date range filter
    try:
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            start = timezone.make_aware(start)
            invoices = invoices.filter(created_at__gte=start)

        if end_date:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            end = end.replace(hour=23, minute=59, second=59)
            end = timezone.make_aware(end)
            invoices = invoices.filter(created_at__lte=end)
    except ValueError:
        pass

    # Calculate total sum of total_amount from filtered invoices
    total_sum = invoices.aggregate(total=Sum('total_amount'))['total'] or 0

    # Pagination
    paginator = Paginator(invoices, 15)
    page_number = request.GET.get('page')
    invoices_page = paginator.get_page(page_number)

    return render(request, 'billing/invoice_list.html', {
        'invoices': invoices_page,
        'total_sum': total_sum,
        'query': query,
        'field': field,
        'start_date': start_date,
        'end_date': end_date,
    })




#===================================================================
from django.shortcuts import render, get_object_or_404
from django.urls import reverse
import qrcode
import base64
from io import BytesIO

from .models import CustomerInvoice, TransactionInvoice, Store

def invoice_detail(request, invoice_number):
    customer_invoice = get_object_or_404(CustomerInvoice, invoice_number=invoice_number)
    
    cart_items = TransactionInvoice.objects.filter(customer_invoice=customer_invoice)
    store_info = Store.objects.get(id=customer_invoice.store_id)

    payment_method_display = customer_invoice.payment_method.capitalize()
    total = sum(item.subtotal for item in cart_items)
    amount_paid = customer_invoice.amount_paid
    change = customer_invoice.change
    user = request.user

    # Get 'next' URL from query parameters or fallback to invoice list page
    next_url = request.GET.get('next', reverse('billing:invoice_list'))

    # QR code generation
    invoice_url = request.build_absolute_uri(
        reverse('billing:invoice_detail', args=[invoice_number])
    )
    qr_img = qrcode.make(invoice_url)
    buffer = BytesIO()
    qr_img.save(buffer, format="PNG")
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()

    context = {
        'customer_invoice': customer_invoice,
        'cart_items': cart_items,
        'store_info': store_info,
        'payment_method_display': payment_method_display,
        'total': total,
        'amount_paid': amount_paid,
        'change': change,
        'user': user,
        'qr_code_base64': qr_code_base64,
        'next_url': next_url,   # Pass next_url to template
    }
    return render(request, 'billing/invoice_receipt.html', context)




#==================================================================
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from django.http import HttpResponse
from collections import defaultdict
from .models import CustomerInvoice


def export_invoices_excel(request):
    # ✅ Exclude voided invoices
    invoices = (
        CustomerInvoice.objects
        .select_related('store', 'user')
        .exclude(is_void=True)  # 👈 exclude voided invoices here
        .order_by('-created_at')
    )

    # --- Filters from request GET params ---
    query = request.GET.get('query')
    field = request.GET.get('field')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if query and field:
        if field == 'invoice_number':
            invoices = invoices.filter(invoice_number__icontains=query)
        elif field == 'user':
            invoices = invoices.filter(user__username__icontains=query)
        elif field == 'store':
            invoices = invoices.filter(store__name__icontains=query)
        elif field == 'customer_name':
            invoices = invoices.filter(customer_name__icontains=query)
        elif field == 'payment_method':
            invoices = invoices.filter(payment_method__icontains=query)

    # --- Date filtering ---
    try:
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            invoices = invoices.filter(created_at__gte=start)
        if end_date:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            end = end.replace(hour=23, minute=59, second=59)
            invoices = invoices.filter(created_at__lte=end)
    except ValueError:
        pass

    # --- Create Workbook ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "Invoice Number", "Cashier", "Store", "Customer Name", "Total Amt",
        "Payment Mtd", "Tax", "Discount", "Final Total", "Amount Paid",
        "Change", "Created At"
    ]
    ws.append(headers)

    # --- Track totals ---
    total_total_amt = total_tax = total_discount = total_final_total = total_amount_paid = total_change = 0

    for invoice in invoices:
        total_amt = float(invoice.total_amount or 0)
        tax = float(invoice.tax or 0)
        discount = float(invoice.discount or 0)
        final_total = float(invoice.final_total or 0)
        amount_paid = float(invoice.amount_paid or 0)
        change = float(invoice.change or 0)

        ws.append([
            invoice.invoice_number or '',
            invoice.user.username if invoice.user else '',
            invoice.store.name if invoice.store else '',
            invoice.customer_name or '',
            total_amt,
            invoice.payment_method or '',
            tax,
            discount,
            final_total,
            amount_paid,
            change,
            invoice.created_at.strftime("%d-%m-%Y %H:%M") if invoice.created_at else ''
        ])

        total_total_amt += total_amt
        total_tax += tax
        total_discount += discount
        total_final_total += final_total
        total_amount_paid += amount_paid
        total_change += change

    # --- Add totals row ---
    total_row = [
        "TOTAL", "", "", "", 
        total_total_amt, "", 
        total_tax, total_discount, 
        total_final_total, total_amount_paid, 
        total_change, ""
    ]
    ws.append(total_row)

    # --- Adjust column widths ---
    for i, column in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        ws.column_dimensions[get_column_letter(i)].width = max(12, max_length + 2)

    # --- Create a summary chart by cashier ---
    user_totals = defaultdict(float)
    for invoice in invoices:
        user = invoice.user.username if invoice.user else 'Unknown'
        final_total = float(invoice.final_total or 0)
        user_totals[user] += final_total

    chart_sheet = wb.create_sheet(title="Totals by Cashier")
    chart_sheet.append(["Cashier", "Total Final Amount"])
    for user, amount in user_totals.items():
        chart_sheet.append([user, amount])

    chart = BarChart()
    chart.title = "Total Sales by Cashier"
    chart.x_axis.title = "Cashier"
    chart.y_axis.title = "Total Final Amount"
    data = Reference(chart_sheet, min_col=2, min_row=1, max_row=chart_sheet.max_row)
    cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=chart_sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    chart_sheet.add_chart(chart, "D2")

    # --- Return Excel file ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Invoices_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


from django.shortcuts import render
from django.db.models import Sum
from django.utils import timezone
from datetime import datetime, time
from .models import CustomerInvoice

def invoice_list_today(request):
    # Get today's date (timezone-aware)
    today = timezone.now()

    # Calculate start and end of the day (timezone-aware)
    start_of_day = datetime.combine(today.date(), time.min)
    end_of_day = datetime.combine(today.date(), time.max)

    # Make sure both start and end of the day are timezone-aware
    start_of_day = timezone.make_aware(start_of_day, timezone.get_current_timezone())
    end_of_day = timezone.make_aware(end_of_day, timezone.get_current_timezone())

    # Check if the logged-in user is an admin or a superuser
    if request.user.is_superuser or request.user.groups.filter(name='admin').exists():
        # Admin users see all invoices
        invoices = CustomerInvoice.objects.filter(
            created_at__gte=start_of_day,
            created_at__lte=end_of_day,
            is_void=False  # Exclude void invoices
        ).order_by('-created_at')
    else:
        # Cashiers only see their own invoices
        invoices = CustomerInvoice.objects.filter(
            created_at__gte=start_of_day,
            created_at__lte=end_of_day,
            is_void=False,  # Exclude void invoices
            user=request.user  # Filter by the currently logged-in user (cashier)
        ).order_by('-created_at')

    # Calculate the total sum of today's invoices (excluding void invoices)
    total_sum = invoices.aggregate(total=Sum('total_amount'))['total'] or 0

    # Pass invoices and total sum to the template
    context = {
        'invoices': invoices,
        'total_sum': total_sum,
    }
    return render(request, 'billing/invoice_list.html', context)


import csv
import openpyxl
from django.http import HttpResponse
from .models import Customer  # Adjust if your model is elsewhere

def export_customers_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=customers.csv'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Name', 'Phone Number'])

    for c in Customer.objects.all():
        writer.writerow([c.id, c.name, c.phone_number or '—'])

    return response


def export_customers_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    ws.append(['ID', 'Name', 'Phone Number'])

    for c in Customer.objects.all():
        ws.append([c.id, c.name, c.phone_number or '—'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=customers.xlsx'
    wb.save(response)
    return response

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from billing.models import TransactionInvoice
from billing.serializers import TransactionInvoiceSerializer
import logging

logger = logging.getLogger(__name__)

class SyncSalesAPIView(APIView):
    def post(self, request, *args, **kwargs):
        sales_data = request.data.get("sales", [])
        serializer = TransactionInvoiceSerializer(data=sales_data, many=True)

        if serializer.is_valid():
            saved_instances = serializer.save()
            logger.info(f"✅ Saved {len(saved_instances)} transaction(s)")
            return Response(
                {"status": "success", "synced": len(saved_instances)},
                status=status.HTTP_201_CREATED
            )

        logger.error(f"❌ Validation failed: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from billing.models import CustomerInvoice, TransactionInvoice
from store.models import Product

# Check if user is admin
def is_admin(user):
    return user.is_superuser or user.is_staff

@login_required
@user_passes_test(is_admin)
def void_invoice(request, invoice_id):
    invoice = get_object_or_404(CustomerInvoice, id=invoice_id)
    
    # Void the customer invoice
    invoice.is_void = True
    invoice.save()

    # Retrieve all transactions linked to that invoice
    transactions = TransactionInvoice.objects.filter(customer_invoice=invoice)

    # Roll back stock and void transactions
    for t in transactions:
        # Add quantity back to product stock
        if t.product and t.quantity:
            t.product.quantity += t.quantity
            t.product.save()

        # Mark the transaction as void
        t.is_void = True
        t.save()

    messages.success(request, f"Invoice #{invoice.id} and its transactions have been voided successfully, and stock has been restored.")
    return redirect('billing:invoice_list')


@login_required
@user_passes_test(is_admin)
def voided_invoices(request):
    # Fetch all invoices that have been voided
    voided_list = CustomerInvoice.objects.filter(is_void=True).order_by('-created_at')

    context = {
        'voided_list': voided_list,
    }
    return render(request, 'billing/voided_invoices.html', context)
