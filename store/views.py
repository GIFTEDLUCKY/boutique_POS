from django.shortcuts import render, get_object_or_404, redirect
from .models import Store, Product, Category, Supplier, Staff
from .forms import ProductForm, StoreForm, CategoryForm, SupplierForm, StaffForm
from django.contrib import messages

from accounts.decorators import role_required
from django.shortcuts import render

# @role_required(allowed_roles=['staff', 'admin'])
# def staff_view(request):
#     return render(request, 'store/staff.html')

# @role_required(allowed_roles=['staff', 'admin'])
# def edit_supplier(request):
#     return render(request, 'store/edit_supplier.html')

# @role_required(allowed_roles=['staff', 'admin'])
# def add_supplier(request):
#     return render(request, 'store/add_supplier.html')

# @role_required(allowed_roles=['staff', 'admin'])
def add_store(request):
    return render(request, 'store/add_store.html')



# View for displaying all stores
def store_list(request):
    stores = Store.objects.all()
    print(stores)
    return render(request, 'store/store_list.html', {'stores': stores})

# View for displaying products, optionally filtered by store
def product_list(request, store_id=None):
    if store_id:
        store = get_object_or_404(Store, id=store_id)
        products = Product.objects.filter(store=store)
    else:
        if hasattr(request.user, 'staff') and hasattr(request.user.staff, 'store'):
            store = request.user.staff.store
            products = Product.objects.filter(store=store)
        else:
            store = None
            products = Product.objects.none()

    return render(request, 'store/product_list.html', {'store': store, 'products': products})




# View for adding a new product
from django.shortcuts import render, redirect
from store.models import Product
from store.forms import ProductForm

import csv
from django.http import HttpResponse

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from django.contrib.auth.decorators import login_required

@login_required
def add_product(request):
     # Export logic
     if 'export' in request.GET:
         search_field = request.GET.get('search_field', '')
         search_value = request.GET.get('search_value', '')

         allowed_fields = ['name', 'category', 'supplier', 'store', 'status']
         products = Product.objects.all()

         if search_field in allowed_fields and search_value:
             if search_field == 'category':
                 products = products.filter(category__name__icontains=search_value)
             elif search_field == 'supplier':
                 products = products.filter(supplier__supplier_name__icontains=search_value)
             elif search_field == 'store':
                 products = products.filter(store__name__icontains=search_value)
             elif search_field == 'status':
                 if search_value.lower() == 'active':
                     products = products.filter(status=True)
                 elif search_value.lower() == 'inactive':
                     products = products.filter(status=False)
             else:
                 products = products.filter(**{f'{search_field}__icontains': search_value})

         # Create Excel file
         wb = openpyxl.Workbook()
         ws = wb.active
         ws.title = "Products"

         # Add headers
         headers = [
             "Name", "Category", "Supplier", "Store", "Quantity",
             "Cost Price", "Selling Price", "Discount", "Final Price", "Assumed Profit", "Status"
         ]
         for col_num, header in enumerate(headers, start=1):
             ws.cell(row=1, column=col_num, value=header)

         # Add product data
         for row_num, product in enumerate(products, start=2):
            tax = (product.selling_price * product.product_tax) / 100
            final_price = (product.selling_price - product.discount) + tax

            ws.cell(row=row_num, column=1, value=product.name)
            ws.cell(row=row_num, column=2, value=product.category.name)
            ws.cell(row=row_num, column=3, value=product.supplier.supplier_name)
            ws.cell(row=row_num, column=4, value=product.store.name)
            ws.cell(row=row_num, column=5, value=product.quantity)
            ws.cell(row=row_num, column=6, value=product.cost_price)
            ws.cell(row=row_num, column=7, value=product.selling_price)
            ws.cell(row=row_num, column=8, value=product.discount)
            ws.cell(row=row_num, column=9, value=final_price)  # Updated to Final Price
            ws.cell(row=row_num, column=10, value=product.assumed_profit)
            ws.cell(row=row_num, column=11, value="Active" if product.status else "Inactive")

         # Set column widths
         for col_num, _ in enumerate(headers, start=1):
             ws.column_dimensions[get_column_letter(col_num)].width = 15

         # Prepare the response
         response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
         response['Content-Disposition'] = 'attachment; filename=products.xlsx'
         wb.save(response)
         return response

     # Handle the search functionality
     search_field = request.GET.get('search_field', '')
     search_value = request.GET.get('search_value', '')

     allowed_fields = ['name', 'category', 'supplier', 'store', 'status']
     products = Product.objects.all()

     if search_field in allowed_fields and search_value:
         if search_field == 'category':
             products = products.filter(category__name__icontains=search_value)
         elif search_field == 'supplier':
             products = products.filter(supplier__supplier_name__icontains=search_value)
         elif search_field == 'store':
             products = products.filter(store__name__icontains=search_value)
         elif search_field == 'status':
             if search_value.lower() == 'active':
                 products = products.filter(status=True)
             elif search_value.lower() == 'inactive':
                 products = products.filter(status=False)
         else:
             products = products.filter(**{f'{search_field}__icontains': search_value})

     # Handle form submission for adding products
     if request.method == 'POST':
         form = ProductForm(request.POST)
         if form.is_valid():
             form.save()
             messages.success(request, "Product added successfully!")  # ✅ Success message
             return redirect('store:add_product')
         else:
             messages.error(request, "Error adding product. Please check the form.")  # ❌ Error message
     else:
         form = ProductForm()

     # Calculate Final Price for each product
     for product in products:
        tax = (product.selling_price * product.product_tax) / 100  # Calculates percentage
        product.final_price = (product.selling_price - product.discount) + tax


     # Return the response with search parameters and filtered products
     return render(request, 'store/add_product.html', {
         'form': form,
         'products': products,
         'search_field': search_field,
         'search_value': search_value
     })


from django.http import JsonResponse
from .models import Product

def scan_barcode(request):
    if request.method == "POST":
        barcode_data = request.POST.get('barcode', '')
        try:
            # Search for the product by barcode
            product = Product.objects.get(barcode=barcode_data)
            response_data = {
                'product_name': product.name,
                'price': product.price,
                'quantity': product.quantity,
            }
            return JsonResponse(response_data)
        except Product.DoesNotExist:
            return JsonResponse({'error': 'Product not found'}, status=404)
    return JsonResponse({'error': 'Invalid request'}, status=400)

from django.http import JsonResponse
from .models import Product




def handle_barcode(request):
    barcode = request.GET.get('barcode')
    if not barcode:
        return JsonResponse({'error': 'No barcode provided'}, status=400)

    try:
        # Check if the product with this barcode exists
        product = Product.objects.get(barcode=barcode)
        
        # If product exists, update its stock or perform any other necessary action
        product.stock += 1  # Example: Increase stock
        product.save()
        return JsonResponse({'message': f'Product {product.name} updated successfully.', 'product_id': product.id})

    except Product.DoesNotExist:
        # If product doesn't exist, create a new product
        product = Product.objects.create(
            barcode=barcode,
            name="New Product",  # Set default values or request user input for these
            price=0.00,  # Default price or fetch from elsewhere
            stock=1  # Default stock value
        )
        return JsonResponse({'message': 'New product added to the database.', 'product_id': product.id})


# Edit Product View
def edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    categories = Category.objects.all()  # Fetch all categories

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()  # Update the product
            return redirect('store:add_product')  # Redirect to the add_product page after saving
    else:
        form = ProductForm(instance=product)

    # Fetch all products to render the updated table
    products = Product.objects.all()

    return render(request, 'store/add_product.html', {
        'form': form,
        'categories': categories,
        'product': product,
        'products': products  # Ensure that the updated list of products is passed to the template
    })


# Delete Product View
def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.delete()  # Delete the product
    return redirect('store:add_product')  # Redirect to the add_product page after deletion


#=========================================================
# VIEWS FOR ADDING STAFF
# View for adding a new staff member
from django.shortcuts import render, redirect
from .forms import StaffForm
from .models import Staff
from django.db.models import Q  # for search functionality

from accounts.models import UserProfile

def add_staff(request):
    staff_members = Staff.objects.all()

    # Search filters (if any search term is entered)
    search_username = request.GET.get('search_username', '')
    search_store = request.GET.get('search_store', '')
    search_role = request.GET.get('search_role', '')

    # Filtering staff based on search terms
    if search_username or search_store or search_role:
        staff_members = staff_members.filter(
            Q(user__username__icontains=search_username) &
            Q(store__name__icontains=search_store) &
            Q(role__icontains=search_role)
        )
    
    if request.method == 'POST':
        form = StaffForm(request.POST)
        if form.is_valid():
            staff = form.save()  # Save the staff member

            # Now ensure the UserProfile is updated with store and role
            user_profile, created = UserProfile.objects.get_or_create(user=staff.user)  # Get or create the profile
            user_profile.store = form.cleaned_data['store']  # Update store
            user_profile.role = form.cleaned_data['role']  # Update role
            user_profile.save()  # Save the profile with updates

            return redirect('store:add_staff')  # Reload the page to show updated list
    else:
        form = StaffForm()  # Show empty form

    return render(request, 'store/add_staff.html', {
        'form': form,
        'staff_list': staff_members,  # Pass the filtered staff list
    })


from django.shortcuts import render, get_object_or_404
from .models import Staff

from django.shortcuts import get_object_or_404, redirect, render
from .forms import StaffForm  # Ensure this is the correct form
from .models import Staff


def edit_staff(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)
    
    if request.method == 'POST':
        form = StaffForm(request.POST, instance=staff)
        if form.is_valid():
            form.save()
            messages.success(request, "Staff details updated successfully!")
            return redirect('store:add_staff')  # Redirect to staff list
        else:
            messages.error(request, "Error updating staff details. Please check the form.")

    else:
        form = StaffForm(instance=staff)

    return render(request, 'store/edit_staff.html', {'form': form, 'staff': staff})

# View for displaying staff members of a store
def staff_list(request):
    staff_members = Staff.objects.filter(store=request.user.staff.store)
    return render(request, 'store/staff_list.html', {'staff_members': staff_members})

from django.shortcuts import render, get_object_or_404, redirect
from .models import Staff

from django.shortcuts import get_object_or_404, redirect
from .models import Staff

from django.contrib import messages

def delete_staff(request, id):
    staff_member = Staff.objects.filter(id=id).first()
    if not staff_member:
        messages.error(request, "Staff member not found.")
        return redirect('store:add_staff')
    staff_member.delete()
    messages.success(request, "Staff member deleted successfully.")
    return redirect('store:add_staff')




# View for displaying categories
def category_list(request):
    categories = Category.objects.all()  # Fetch all categories
    return render(request, 'store/category_list.html', {'categories': categories})

# View for adding a new supplier


# Add Supplier View
from django.shortcuts import render, redirect
from .models import Supplier
from .forms import SupplierForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages

@login_required
def add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Supplier added successfully!")  # ✅ Success message
            return redirect('store:add_supplier')  # Redirect to the same page after saving
        else:
            messages.error(request, "Error adding supplier. Please check the form.")  # ❌ Error message

    else:
        form = SupplierForm()

    # Fetch updated supplier list
    suppliers = Supplier.objects.all()

    # Optional: Add search functionality
    search_invoice_no = request.GET.get('search_invoice_no', '')
    search_supplier_name = request.GET.get('search_supplier_name', '')
    search_supplier_contact = request.GET.get('search_supplier_contact', '')

    if search_invoice_no:
        suppliers = suppliers.filter(invoice_no__icontains=search_invoice_no)
    if search_supplier_name:
        suppliers = suppliers.filter(supplier_name__icontains=search_supplier_name)
    if search_supplier_contact:
        suppliers = suppliers.filter(supplier_contact__icontains=search_supplier_contact)

    context = {
        'form': form,
        'suppliers': suppliers,
    }
    return render(request, 'store/add_supplier.html', context)


# Edit Supplier View
def edit_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, "Supplier updated successfully!")  # ✅ Success message
            return redirect('store:add_supplier')  # Redirect to the supplier list page
        else:
            messages.error(request, "Error updating supplier. Please check the form.")  # ❌ Error message

    else:
        form = SupplierForm(instance=supplier)

    context = {
        'form': form,
        'supplier': supplier,
    }
    return render(request, 'store/edit_supplier.html', context)



from django.db.models import ProtectedError

@login_required
def delete_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    
    if request.method == 'POST':
        try:
            supplier.delete()
            messages.success(request, "Supplier deleted successfully!")
        except ProtectedError as e:
            # protected_objects is a set, so iterate directly
            protected_products = e.protected_objects
            product_names = ", ".join([str(p) for p in protected_products])
            user_message = (
                f"Cannot delete this supplier because it is associated with these products: {product_names}. "
                "Please reassign or delete those products first."
            )
            messages.error(request, user_message)
        except Exception as e:
            messages.error(request, f"Error deleting supplier: {e}")
        
        return redirect('store:add_supplier')

    context = {
        'supplier': supplier,
    }
    return render(request, 'store/confirm_delete.html', context)



@login_required
# View for adding a new category
def add_category(request):
    categories = Category.objects.all()

    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Category added successfully!")  # ✅ Success message
            return redirect('store:add_category')  # Redirect to the same page after form submission
        else:
            messages.error(request, "Error adding category. Please check the form.")  # ❌ Error message

    else:
        form = CategoryForm()

    return render(request, 'store/add_category.html', {'form': form, 'categories': categories})

from django.http import HttpResponse

# Edit Category View
@login_required
def edit_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)

    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Category updated successfully!")  # ✅ Success message
            return redirect('store:add_category')  # Redirect back to the category list page
        else:
            messages.error(request, "Error updating category. Please check the form.")  # ❌ Error message

    else:
        form = CategoryForm(instance=category)

    return render(request, 'store/add_category.html', {
        'form': form,
        'categories': Category.objects.all(),
    })


#===================================================================
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import ProtectedError
from .models import Category

def delete_category(request, category_id):  # or use pk if your URL uses that param name
    category = get_object_or_404(Category, pk=category_id)

    if request.method == "POST":
        try:
            category.delete()
            messages.success(request, "Category deleted successfully!")
            return redirect('store:add_category')  # Redirect to your category list page/view
        except ProtectedError as e:
            protected_objects = e.protected_objects
            product_names = ", ".join(str(p) for p in protected_objects)
            user_message = (
                f"Cannot delete this category because it is associated with these products: {product_names}. "
                "Please reassign or delete those products first."
            )
            messages.error(request, user_message)
            return redirect('store:add_category')
        except Exception as e:
            messages.error(request, f"Error deleting category: {str(e)}")
            return redirect('store:add_category')

    return render(request, 'store/confirm_delete_category.html', {'category': category})


#================================================================
def edit_store(request, pk=None):
    store = get_object_or_404(Store, pk=pk) if pk else None
    form = StoreForm(request.POST or None, instance=store)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            if pk:
                messages.success(request, "Store updated successfully!")
            else:
                messages.success(request, "New store added successfully!")
            return redirect('store:add_store')
        else:
            messages.error(request, "Error processing the form. Please check your inputs.")

    stores = Store.objects.all()
    return render(request, 'store/add_store.html', {'form': form, 'stores': stores, 'store': store})


#====================================================================
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import ProtectedError
from .models import Store
from django.contrib.auth.decorators import login_required

@login_required
def delete_store(request, pk):
    store = get_object_or_404(Store, pk=pk)
    
    if request.method == 'POST':
        try:
            store.delete()
            messages.success(request, "Store deleted successfully!")
            return redirect('store:add_store')  # Redirect to your store list page
        except ProtectedError as e:
            protected_products = e.protected_objects
            product_names = ", ".join(str(p) for p in protected_products)
            error_msg = (
                f"Cannot delete this store because it is associated with these products: {product_names}. "
                "Please reassign or delete those products first."
            )
            messages.error(request, error_msg)
            return redirect('store:add_store')  # Redirect back to store list with error message
        except Exception as e:
            messages.error(request, f"Error deleting store: {e}")
            return redirect('store:add_store')

    context = {
        'store': store,
    }
    return render(request, 'store/confirm_delete_store.html', context)

#===================================================================
from django.shortcuts import render
from django.db.models import Q
from .models import Product  # Adjust as needed

def search_product(request):
    products = Product.objects.all()  # Show all products by default

    if request.method == "POST":
        search_field = request.POST.get("search_field")
        search_value = request.POST.get("search_value")

        if search_field and search_value:
            if search_field == "name":
                products = products.filter(name__icontains=search_value)
            elif search_field == "category":
                products = products.filter(category__icontains=search_value)
            elif search_field == "supplier":
                products = products.filter(supplier__supplier_name__icontains=search_value)
            elif search_field == "store":
                products = products.filter(store__icontains=search_value)
            elif search_field == "status":
                products = products.filter(status__icontains=search_value)

    return render(request, 'store/add_product.html', {"products": products})

from django.http import HttpResponse
from openpyxl import Workbook
from .models import Product
from django.http import HttpResponse
import xlwt  # Install xlwt library for writing Excel files

def export_to_excel(request):
    # Get search filters from the request
    search_field = request.GET.get('search_field', '')
    search_value = request.GET.get('search_value', '')

    # Apply filters if search parameters are provided
    products = Product.objects.all()
    if search_field and search_value:
        if search_field == 'status':
            status_value = True if search_value == 'Active' else False
            products = products.filter(status=status_value)
        elif search_field == 'supplier':
            products = products.filter(supplier__supplier_name__icontains=search_value)
        elif search_field == 'category':
            products = products.filter(category__name__icontains=search_value)
        else:
            products = products.filter(**{f'{search_field}__icontains': search_value})

    # Create Excel workbook and sheet
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('Products')

    # Write headers
    headers = ['ID', 'Name', 'Category', 'Supplier', 'Price', 'Status', 'Store']
    for col_num, header in enumerate(headers):
        sheet.write(0, col_num, header)

    # Write product data
    for row_num, product in enumerate(products, start=1):
        sheet.write(row_num, 0, product.id)
        sheet.write(row_num, 1, product.name)
        sheet.write(row_num, 2, product.category.name if product.category else '')
        sheet.write(row_num, 3, product.supplier.supplier_name if product.supplier else '')
        sheet.write(row_num, 4, product.price)
        sheet.write(row_num, 5, 'Active' if product.status else 'Inactive')
        sheet.write(row_num, 6, product.store.name if product.store else '')

    # Create response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="products.xls"'
    workbook.save(response)
    return response

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from store.models import Product

@login_required
def store_sales(request):
    user_store = request.user.store
    if not user_store:
        return render(request, 'error.html', {'message': 'No store assigned to your account.'})
    
    # Filter products for the logged-in user's store
    products = Product.objects.filter(store=user_store, status=True)

    return render(request, 'billing/sales.html', {'products': products})




# store/views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import Store
from .forms import StoreForm
from django.contrib.auth.decorators import login_required


@login_required
def add_store(request, store_id=None):
    store = get_object_or_404(Store, id=store_id) if store_id else None

    if request.method == 'POST':
        form = StoreForm(request.POST, instance=store)
        if form.is_valid():
            if store:
                messages.success(request, "Store updated successfully!")
            else:
                messages.success(request, "New store added successfully!")
            form.save()
            return redirect('store:add_store')  # Redirect to clear the form
        else:
            messages.error(request, "Error saving store. Please check the form.")

    else:
        form = StoreForm(instance=store)

    stores = Store.objects.all()

    return render(request, 'store/add_store.html', {
        'form': form,
        'store': store,
        'stores': stores,
    })



from django.shortcuts import render
from .forms import TaxAndDiscountForm
from .models import TaxAndDiscount  # Import the TaxAndDiscount model


def manage_tax_discount(request):
    # Create or fetch the current tax and discount settings
    settings = TaxAndDiscount.objects.first()
    form = TaxAndDiscountForm(instance=settings)

    if request.method == 'POST':
        form = TaxAndDiscountForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, "Tax and discount updated successfully!")
            return redirect('store:manage_tax_discount')  # Reload page after success

    return render(request, 'store/manage_tax_discount.html', {'form': form})


from django.http import JsonResponse
from .models import Product
from django.contrib.auth.decorators import login_required

@login_required
def search_products(request):
    query = request.GET.get('q', '')
    results = []
    if len(query) >= 2:
        products = Product.objects.filter(name__icontains=query, store=request.user.store)[:10]
        for product in products:
            results.append({
                'id': product.id,
                'name': product.name,
                'price': float(product.discounted_price),
                'stock': product.quantity,
            })
    return JsonResponse({'results': results})
