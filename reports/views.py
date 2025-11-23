import io
import base64
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from django.shortcuts import render
from django.utils.timezone import make_aware
from datetime import datetime
from billing.models import TransactionInvoice
from store.models import Store
from reports.models import PriceHistory

def make_autopct(values):
    def my_autopct(pct):
        total = sum(values)
        val = int(round(pct * total / 100.0))
        return f'{val}'
    return my_autopct


def generate_bar_chart(labels, values, title, xlabel="Category", ylabel="Amount"):
    values = [float(v) for v in values]  # ensure floats

    fig, ax = plt.subplots(figsize=(8, 6))
    bars = ax.bar(labels, values, color=['green' if v >= 0 else 'red' for v in values])

    ax.set_title(title)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)

    # Add value labels on top of bars
    for bar, value in zip(bars, values):
        height = bar.get_height()
        ax.annotate(
            f'{value:.2f}',
            xy=(bar.get_x() + bar.get_width() / 2, height),
            xytext=(0, 3),  # 3 points vertical offset
            textcoords="offset points",
            ha='center',
            va='bottom'
        )

    plt.xticks(rotation=45, ha='right')

    buf = io.BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format='png', bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    image_png = buf.getvalue()
    buf.close()
    return base64.b64encode(image_png).decode('utf-8')


def profit_and_loss_view(request):
    store_id = request.GET.get('store')
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    # ✅ Exclude voided invoices
    invoices = TransactionInvoice.objects.select_related('product', 'store', 'customer_invoice') \
        .filter(customer_invoice__is_void=False)

    selected_store = None

    if store_id is not None and store_id.isdigit():
        invoices = invoices.filter(product__store_id=store_id)
        selected_store = int(store_id)

    start_dt = None
    end_dt = None

    if start_date and end_date:
        try:
            start_dt = make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
            end_dt = make_aware(datetime.strptime(end_date, "%Y-%m-%d"))
            if start_dt.date() == end_dt.date():
                end_dt = end_dt.replace(hour=23, minute=59, second=59)
            invoices = invoices.filter(created_at__range=(start_dt, end_dt))
        except ValueError:
            pass

    # Get stores based on filtered invoices' product stores only
    stores = Store.objects.filter(id__in=invoices.values_list('product__store_id', flat=True).distinct())

    summary = {}
    for invoice in invoices:
        product = invoice.product
        store = product.store

        qty = invoice.quantity
        unit_cost = product.cost_price
        unit_final_price = invoice.adjusted_final_price / invoice.quantity if invoice.quantity else 0


        total_cost = unit_cost * qty
        total_revenue = unit_final_price * qty
        profit = total_revenue - total_cost

        key = (product.id, store.id)
        if key not in summary:
            summary[key] = {
                'product_name': product.name,
                'store_name': store.name,
                'total_qty': 0,
                'unit_cost': unit_cost,
                'unit_price': invoice.price,  # price before discounts/tax (unit)
                'final_selling_price': unit_final_price,  # unit final price (after tax & discount)
                'total_cost': 0,
                'total_revenue': 0,
                'total_profit': 0,
            }

        summary[key]['total_qty'] += qty
        summary[key]['total_cost'] += total_cost
        summary[key]['total_revenue'] += total_revenue
        summary[key]['total_profit'] += profit


    # Calculate unit cost and profit for each summary item
    for (product_id, store_id), item in summary.items():
        product_name = item['product_name']
        store_name = item['store_name']

        if start_dt:
            latest_price = PriceHistory.objects.filter(
                product__name=product_name,
                product__store__name=store_name,
                date_changed__lte=start_dt
            ).order_by('-date_changed').first()
        else:
            latest_price = None

        if latest_price:
            unit_cost = latest_price.new_cp
        else:
            invoice_instance = TransactionInvoice.objects.filter(
                product__name=product_name,
                product__store__name=store_name
            ).first()

            if invoice_instance and invoice_instance.product:
                unit_cost = invoice_instance.product.cost_price
            else:
                unit_cost = 0

        item['unit_cost'] = unit_cost
        item['total_cost'] = round(unit_cost * item['total_qty'], 2)
        item['total_profit'] = round(item['total_revenue'] - item['total_cost'], 2)

    total_revenue = sum(item['total_revenue'] for item in summary.values())
    total_cost = sum(item['total_cost'] for item in summary.values())
    total_profit = sum(item['total_profit'] for item in summary.values() if item['total_profit'] > 0)
    total_loss = sum(item['total_profit'] for item in summary.values() if item['total_profit'] < 0)


    # Prepare data for pie charts:
    # Profit by Store
    profit_by_store = {}
    for item in summary.values():
        store_name = item['store_name']
        profit_by_store[store_name] = profit_by_store.get(store_name, 0) + item['total_profit']

    # Remove negatives
    profit_by_store = {k: v for k, v in profit_by_store.items() if v > 0}

    # Profit by Product
    profit_by_product = {}
    for item in summary.values():
        product_name = item['product_name']
        profit_by_product[product_name] = profit_by_product.get(product_name, 0) + item['total_profit']

    profit_by_product = {k: v for k, v in profit_by_product.items() if v > 0}

    # Generate pie charts images in base64
    store_bar_chart = None   # ✅ always define a default
    if profit_by_store:
        store_bar_chart = generate_bar_chart(
        labels=list(profit_by_store.keys()),
        values=list(profit_by_store.values()),
        title="Profit/Loss by Store",
        xlabel="Store",
        ylabel="Profit / Loss"
    )


    product_bar_chart = None # ✅ always define a default
    if profit_by_product:
        product_bar_chart = generate_bar_chart(
        labels=list(profit_by_product.keys()),
        values=list(profit_by_product.values()),
        title="Profit/Loss by Product",
        xlabel="Product",
        ylabel="Profit / Loss"
    )

    context = {
        'invoices': list(summary.values()),
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'total_loss': total_loss,  # ✅ ✅ ✅ You forgot this!
        'stores': stores,
        'selected_store': selected_store,
        'start_date': start_date,
        'end_date': end_date,
        'store_bar_chart': store_bar_chart,
        'product_bar_chart': product_bar_chart,

    }

    return render(request, 'reports/profit_and_loss.html', context)



#===========================================================
# reports/views.py



from django.db.models import Sum, F, FloatField, ExpressionWrapper
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from billing.models import TransactionInvoice
from store.models import Store

def export_profit_loss_excel(request):
    def clean_param(value):
        return value if value and value != 'None' else None

    store_id = clean_param(request.GET.get('store'))
    start_date = clean_param(request.GET.get('start'))
    end_date = clean_param(request.GET.get('end'))

    txns = TransactionInvoice.objects.select_related('product', 'store', 'customer_invoice')

    if store_id:
        txns = txns.filter(product__store_id=store_id)
    if start_date:
        txns = txns.filter(created_at__date__gte=start_date)
    if end_date:
        txns = txns.filter(created_at__date__lte=end_date)


    # Annotate final selling price (price - discount per unit)
    txns = txns.annotate(
        final_price=ExpressionWrapper(
            F('price') - F('discount'),
            output_field=FloatField()
        ),
        total_cost=ExpressionWrapper(
            F('quantity') * F('product__cost_price'),
            output_field=FloatField()
        ),
        total_revenue=ExpressionWrapper(
            F('quantity') * (F('price') - F('discount')),
            output_field=FloatField()
        )
    )

    # Group by store & product
    summary = {}
    for txn in txns:
        key = (txn.store.name if txn.store else 'N/A', txn.product.name)
        if key not in summary:
            summary[key] = {
                'store': txn.store.name if txn.store else 'N/A',
                'product': txn.product.name,
                'quantity': 0,
                'unit_cost': float(txn.product.cost_price),
                'unit_price': float(txn.price),
                'final_price': float(txn.final_price),
                'total_cost': 0,
                'total_revenue': 0
            }

        summary[key]['quantity'] += txn.quantity
        summary[key]['total_cost'] += txn.quantity * float(txn.product.cost_price)
        summary[key]['total_revenue'] += txn.quantity * float(txn.final_price)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit and Loss"

    # Compute grand totals
    total_cost = sum(item['total_cost'] for item in summary.values())
    total_revenue = sum(item['total_revenue'] for item in summary.values())
    total_profit = total_revenue - total_cost

    # Write summary at top
    ws.append([f"Total Revenue ₦: {total_revenue:.2f}"])
    ws.append([f"Total Cost ₦: {total_cost:.2f}"])
    ws.append([f"Profit ₦: {total_profit:.2f}"])
    ws.append([])  # empty row

    # Write table headers
    headers = [
        'S/NO.', 'Store', 'Product', 'Total Qty', 'Unit Cost', 'Unit Price',
        'Final Selling Price', 'Total Cost', 'Total Revenue', 'Total Profit'
    ]
    ws.append(headers)

    for idx, ((store, product), data) in enumerate(summary.items(), start=1):
        total_profit = data['total_revenue'] - data['total_cost']
        ws.append([
            idx,
            data['store'],
            data['product'],
            data['quantity'],
            f"{data['unit_cost']:.2f}",
            f"{data['unit_price']:.2f}",
            f"{data['final_price']:.2f}",
            f"{data['total_cost']:.2f}",
            f"{data['total_revenue']:.2f}",
            f"{total_profit:.2f}"
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Return response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = "profit_loss_summary.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


from django.shortcuts import render, redirect
from reports.forms import PriceHistoryForm
from django.contrib.auth.decorators import login_required
from django.db import transaction, IntegrityError
from django.contrib import messages
from reports.models import PriceHistory
import logging

logger = logging.getLogger(__name__)

@login_required
def price_history_view(request):
    if request.method == 'POST':
        form = PriceHistoryForm(request.POST)
        if form.is_valid():
            product = form.cleaned_data.get('product')
            new_cp = form.cleaned_data.get('new_cp')
            new_sp = form.cleaned_data.get('new_sp')

            # --- Validate required fields ---
            missing_fields = []
            if product is None:
                missing_fields.append("Product")
            if new_cp is None:
                missing_fields.append("New Cost Price")
            if new_sp is None:
                missing_fields.append("New Selling Price")

            if missing_fields:
                messages.error(request, f"Please fill in the following fields: {', '.join(missing_fields)}")
            else:
                try:
                    with transaction.atomic():
                        # Save old prices first
                        old_cp = product.cost_price
                        old_sp = product.selling_price

                        # --- Update product once ---
                        product.cost_price = new_cp
                        product.selling_price = new_sp
                        product.save(update_fields=['cost_price', 'selling_price'])

                        # --- Create PriceHistory record with changed_by ---
                        PriceHistory.objects.create(
                            product=product,
                            old_cp=old_cp,
                            new_cp=new_cp,
                            old_sp=old_sp,
                            new_sp=new_sp,
                            final_price=new_sp,  # adjust if needed
                            changed_by=request.user
                        )

                        logger.info(f"Updated Product {product.id} prices to CP={product.cost_price}, SP={product.selling_price}")
                        messages.success(request, f"Prices for '{product.name}' updated successfully.")

                    return redirect('price_history')

                except IntegrityError as e:
                    logger.error(f"Database error while updating prices: {e}")
                    messages.error(request, "Database error: Could not save price history. Please check your inputs.")
        else:
            # Add form field errors to messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    else:
        form = PriceHistoryForm()

    # --- GET: filter/search PriceHistory ---
    history = PriceHistory.objects.select_related('product__store', 'changed_by').all().order_by('-date_changed')

    store = request.GET.get('store', '').strip()
    product_name = request.GET.get('product', '').strip()
    user = request.GET.get('user', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()

    if store:
        history = history.filter(product__store__name__icontains=store)
    if product_name:
        history = history.filter(product__name__icontains=product_name)
    if user:
        history = history.filter(changed_by__username__icontains=user)
    if start_date:
        history = history.filter(date_changed__date__gte=start_date)
    if end_date:
        history = history.filter(date_changed__date__lte=end_date)

    return render(request, 'reports/price_history.html', {
        'form': form,
        'history': history,
        'request': request,
        'total_records': history.count()
    })




from django.http import JsonResponse
from store.models import Product

def product_price_api(request):
    product_id = request.GET.get('product_id')
    data = {'old_cp': '', 'old_sp': ''}
    if product_id:
        try:
            product = Product.objects.get(id=product_id)
            data['old_cp'] = str(product.cost_price)
            data['old_sp'] = str(product.selling_price)
        except Product.DoesNotExist:
            pass
    return JsonResponse(data)

